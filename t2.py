# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from itertools import islice
import os
import sys
import pandas as pd
import numpy as np


start = '2018-01-01'
end = '2018-12-31'


def GetInOutDataFrame(path):
  wb = load_workbook(filename=path, read_only=True)
  ws = wb['MOUVEMENTS']

  expected_initial_header = 'Nom'
  header_row_index = 3
  row_index = header_row_index
  if ws.cell(row_index, 1).value != expected_initial_header:
    raise NameError('La valeur attendue de la cellule ' + ws.cell(row_index, 1).coordinate + ' est "' + expected_initial_header + '".')
  while ws.cell(row_index, 1).value:
    row_index = row_index + 1

  column_limit = 'Observations cartographie'
  column_max = 20
  column_index = 1
  columns = []
  while ws.cell(header_row_index, column_index).value != column_limit and column_index < column_max:
    columns.append(ws.cell(header_row_index, column_index).value)
    column_index = column_index + 1
  columns.append(ws.cell(header_row_index, column_index).value)

  if column_index == column_max:
    raise NameError('Il n\'apparait pas une colonne "' + column_limit + '" dans les 20 premières colonnes.')

  data = list(ws.values)[header_row_index:]
  raw_df = pd.DataFrame(data, columns=columns)

  columns = ['Nom', u'Prénom', u'Date d\'entrée', 'Date de sortie', column_limit]
  new_columns = ['NOM', 'PRENOM', 'ENTREE', 'SORTIE', 'IMPUTATION']
  df = raw_df[columns]

  df.columns = new_columns

  df = df[-df.ENTREE.isnull()].copy()

  for col in df.columns:
    if df[col].dtype == 'object':
      df[col] = df[col].str.strip().str.upper()

  df.loc[df.SORTIE.isnull(), 'SORTIE'] = end
  df['MATCH'] = 1

  df.ENTREE = pd.to_datetime(df.ENTREE)
  df.SORTIE = pd.to_datetime(df.SORTIE)

  return df


def getDailyImputation(df):
  days = pd.date_range(start=start, end=end, freq='B')
  day_df = pd.DataFrame({ 'MATCH': 1, 'JOUR':days})

  daily_df = df.merge(day_df, on='MATCH')
  idx = (daily_df.ENTREE <= daily_df.JOUR) & (daily_df.JOUR <= daily_df.SORTIE)
  filterd_df = daily_df[idx].copy()

  return filterd_df.set_index(['NOM', 'PRENOM', 'JOUR'])[['IMPUTATION']]


def getT2(path):
  wb = load_workbook(filename=path, read_only=True)
  ws = wb['Feuil1']
  data = ws.values
  columns = next(data)
  data = list(data)
  raw_df = pd.DataFrame(data, columns=columns)

  period_columns = [i for i in range(13) if i in raw_df.columns]
  if not len(period_columns):
    raise KeyError('Aucune période n\'est présente dans le fichier.', raw_df.columns)

  required_columns = ['NOM', 'PRENOM']
  preserved_columns = required_columns + period_columns

  missing = [required for required in required_columns if required not in preserved_columns]
  if len(missing):
    raise KeyError('Des champs obligatoires sont manquants : ' + missing)

  limited_df = raw_df[preserved_columns]
  for col in limited_df.columns:
    if limited_df[col].dtype == 'object':
      limited_df.loc[:,col] = limited_df[col].str.strip().str.upper()

  long_df = limited_df.melt(id_vars=required_columns, var_name='PERIODE', value_name='MONTANT')
  long_df = (long_df[-long_df.MONTANT.isnull()]).copy()

  return long_df.groupby([] + required_columns + ['PERIODE']).sum()


"""
  NOM, PRENOM, JOUR -> IMPUTATION
  NOM, PRENOM, PERIOD, MONTANT

  NOM, PRENOM, JOUR, IMPUTATION, MONTANT
"""
def getDailyAttribution(daily_moves, t2):
  # Add PERIOD to join with T2
  daily_moves.loc[:,'PERIODE'] = daily_moves.reset_index().JOUR.dt.month.values
  indexes = ['NOM', 'PRENOM', 'PERIODE']

  # Count rows to spread MONTANT from T2
  day_counts = pd.Series(daily_moves.groupby(indexes).IMPUTATION.count(), name='NOMBREJOUR')

  t2_period = pd.concat([t2.reset_index().set_index(indexes), day_counts], axis=1, join='outer')
  t2_period.MONTANT[t2_period.MONTANT.isnull()] = 0
  t2_period.NOMBREJOUR[t2_period.NOMBREJOUR.isnull()] = 1

  t2_period = t2_period.assign(MONTANTPERIODE=t2_period.MONTANT/t2_period.NOMBREJOUR)

  result = daily_moves.merge(pd.DataFrame(t2_period.MONTANTPERIODE), on=indexes, how='outer')

  if not np.isclose(result.MONTANTPERIODE.sum(), t2.MONTANT.sum()):
    raise ValueError('Erreur de manipulation', result.MONTANTPERIODE.sum(), t2.MONTANT.sum())

  return result


def main():
  mouvements_path = '/home/thomas/Documents/Beta.gouv.fr/RH/2018-11-29'

  moves_dinsic = GetInOutDataFrame(os.path.join(mouvements_path, 'MOUVEMENTS DINSIC.xlsm'))
  moves_rie = GetInOutDataFrame(os.path.join(mouvements_path, 'MOUVEMENTS SCNRIE.xlsm'))
  moves = moves_dinsic.append(moves_rie, ignore_index=True)

  daily_moves = getDailyImputation(moves)
  daily_moves.to_csv(os.path.join(mouvements_path, 'moves-test.csv'), encoding='utf-8')

  p = os.path.join(mouvements_path, 'T2.xlsx')
  df = getT2(p)
  df.to_csv(os.path.join(mouvements_path, 't2-test.csv'), encoding='utf-8', sep=";", decimal=",")
  print('T2', df.MONTANT.sum())

  attrib = getDailyAttribution(daily_moves, df)
  attrib.to_csv('attrib-test.csv', encoding='utf-8', sep=";", decimal=",")

  montants = attrib.groupby('IMPUTATION').MONTANTPERIODE.sum()
  unallocated = attrib[attrib.IMPUTATION.isnull()].MONTANTPERIODE.sum()
  montants.to_csv(os.path.join(mouvements_path, 'montants-test.csv'), encoding='utf-8', sep=";", decimal=",")
  print(montants.sum() + unallocated)

  control_groups_df = attrib.groupby(['NOM', 'PRENOM'])
  control_df = pd.DataFrame({ 'MONTANTPERIODE': control_groups_df.MONTANTPERIODE.sum(), 'JOURS': control_groups_df.IMPUTATION.count() })
  control_df.to_csv(os.path.join(mouvements_path, 'control-full-test.csv'), encoding='utf-8', sep=";", decimal=",")

  control_df[(control_df.MONTANTPERIODE == 0) | (control_df.JOURS == 0)].to_csv(os.path.join(mouvements_path, 'control-test.csv'), encoding='utf-8', sep=";", decimal=",")

if __name__ == '__main__':
  sys.exit(main())
